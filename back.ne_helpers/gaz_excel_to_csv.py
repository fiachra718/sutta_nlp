import sys
from pathlib import Path
from openpyxl import load_workbook

src = Path(sys.argv[1])
out = src.with_name("gazetteer_from_bold_rows.txt")

wb = load_workbook(src, data_only=True)
seen, out_lines = set(), []

for ws in wb.worksheets:
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row):
        # treat the row as bold if ANY cell in the row is bold
        if not any(getattr(c.font, "bold", False) for c in r):
            continue
        val = (r[0].value or "").strip()  # Column A
        if val and val not in seen:
            seen.add(val)
            out_lines.append(val)

out.write_text("\n".join(out_lines), encoding="utf-8")
print(f"Wrote {len(out_lines)} lines -> {out}")